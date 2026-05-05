import os
import sqlite3
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import REPORTS_DIR

DB_FILE = "meals.db"

def init_db():
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS meals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            building TEXT,
            stage TEXT,
            grade TEXT,
            litera TEXT,
            class_name TEXT,
            category TEXT,
            quantity INTEGER,
            teacher_name TEXT,
            teacher_id INTEGER,
            created_at TEXT
        )
    """)
    conn.commit()
    conn.close()
    print("✅ База данных инициализирована")

def add_meal(building, stage, grade, litera, class_name, category, quantity, teacher_name, teacher_id):
    date = datetime.now().strftime("%Y-%m-%d")
    created_at = datetime.now().isoformat()
    
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        INSERT INTO meals 
        (date, building, stage, grade, litera, class_name, category, quantity, teacher_name, teacher_id, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (date, building, stage, grade, litera, class_name, category, quantity, teacher_name, teacher_id, created_at))
    conn.commit()
    conn.close()
    print(f"✅ Сохранено: {building} | {stage} | {class_name} | {category} | {quantity}")
    return True

def add_meal_backdated(building, stage, grade, litera, class_name, category, quantity, teacher_name, teacher_id, back_date):
    """Добавляет заявку задним числом"""
    created_at = datetime.now().isoformat()
    
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        INSERT INTO meals 
        (date, building, stage, grade, litera, class_name, category, quantity, teacher_name, teacher_id, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (back_date, building, stage, grade, litera, class_name, category, quantity, teacher_name, teacher_id, created_at))
    conn.commit()
    conn.close()
    print(f"✅ Сохранено задним числом: {back_date} | {building} | {stage} | {class_name} | {category} | {quantity}")
    return True

def has_user_today_request(user_id: int, building: str, class_name: str, stage: str) -> bool:
    today = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.execute("""
        SELECT COUNT(*) FROM meals 
        WHERE teacher_id = ? AND date = ? AND building = ? AND class_name = ? AND stage = ?
    """, (user_id, today, building, class_name, stage))
    count = cursor.fetchone()[0]
    conn.close()
    return count > 0

def has_user_request_on_date(user_id: int, building: str, class_name: str, stage: str, date: str) -> bool:
    """Проверяет, есть ли заявка на конкретную дату"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.execute("""
        SELECT COUNT(*) FROM meals 
        WHERE teacher_id = ? AND date = ? AND building = ? AND class_name = ? AND stage = ?
    """, (user_id, date, building, class_name, stage))
    count = cursor.fetchone()[0]
    conn.close()
    return count > 0

def get_user_meals_today(user_id):
    today = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.execute("""
        SELECT building, class_name, category, quantity, stage
        FROM meals WHERE teacher_id = ? AND date = ?
    """, (user_id, today))
    results = cursor.fetchall()
    conn.close()
    
    meals = []
    for row in results:
        meals.append({
            "building": row[0],
            "class_name": row[1],
            "category": row[2],
            "quantity": row[3],
            "stage": row[4]
        })
    return meals

def get_report_by_building_and_date_range(building: str, date_from: str, date_to: str) -> pd.DataFrame:
    conn = sqlite3.connect(DB_FILE)
    df = pd.read_sql_query("""
        SELECT class_name, category, quantity, stage
        FROM meals 
        WHERE building = ? AND date >= ? AND date <= ? AND stage NOT IN ('after_school', 'home')
        ORDER BY stage, class_name, category
    """, conn, params=(building, date_from, date_to))
    conn.close()
    
    if not df.empty:
        df = df.groupby(['stage', 'class_name', 'category'], as_index=False)['quantity'].sum()
    return df

def get_after_school_requests(building: str, date_from: str, date_to: str) -> pd.DataFrame:
    """Получить заявки на продленку за период (по зданию Марченко или Танкистов)"""
    conn = sqlite3.connect(DB_FILE)
    df = pd.read_sql_query("""
        SELECT class_name, quantity, date
        FROM meals 
        WHERE building = ? AND stage = 'after_school' AND date >= ? AND date <= ?
        ORDER BY date, class_name
    """, conn, params=(building, date_from, date_to))
    conn.close()
    
    if not df.empty:
        df = df.groupby(['date', 'class_name'], as_index=False)['quantity'].sum()
    return df

def get_home_requests(date_from: str, date_to: str) -> pd.DataFrame:
    """Получить заявки надомного отделения за период (только для Марченко)"""
    conn = sqlite3.connect(DB_FILE)
    df = pd.read_sql_query("""
        SELECT class_name, category, quantity
        FROM meals 
        WHERE building = 'Надомное' AND stage = 'home' AND date >= ? AND date <= ?
        ORDER BY class_name, category
    """, conn, params=(date_from, date_to))
    conn.close()
    
    if not df.empty:
        df = df.groupby(['class_name', 'category'], as_index=False)['quantity'].sum()
    return df

def get_shift_by_class(class_name: str) -> int:
    grade = int(class_name.split('.')[0]) if '.' in class_name else 0
    if grade in [1, 4, 5, 7, 9, 10, 11]:
        return 1
    elif grade in [2, 3, 6, 8]:
        return 2
    else:
        return 0

def get_requests_by_shift(building: str, shift: int, date: str = None) -> list:
    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.execute("""
        SELECT class_name, category, quantity, teacher_name
        FROM meals 
        WHERE building = ? AND date = ? AND stage NOT IN ('after_school', 'home')
        ORDER BY class_name, category
    """, (building, date))
    results = cursor.fetchall()
    conn.close()
    
    filtered = []
    for row in results:
        class_name = row[0]
        if get_shift_by_class(class_name) == shift:
            filtered.append({
                "class_name": class_name,
                "category": row[1],
                "quantity": row[2],
                "teacher_name": row[3]
            })
    
    return filtered

def format_requests_by_shift(requests: list, shift: int, building: str, date: str) -> str:
    if not requests:
        return f"📭 Заявок для {building}, {shift} смена на {date} нет"
    
    by_class = {}
    for req in requests:
        class_name = req["class_name"]
        quantity = req["quantity"]
        if class_name not in by_class:
            by_class[class_name] = 0
        by_class[class_name] += quantity
    
    shift_name = "1 смена" if shift == 1 else "2 смена"
    result = f"📋 **Заявки на питание**\n"
    result += f"🏫 **Здание:** {building}\n"
    result += f"📅 **Дата:** {date}\n"
    result += f"👥 **Смена:** {shift_name}\n\n"
    
    total_all = 0
    for class_name in sorted(by_class.keys()):
        total = by_class[class_name]
        result += f"📖 **{class_name} класс:** {total} чел.\n"
        total_all += total
    
    result += f"\n---\n🍽️ **ВСЕГО: {total_all} чел.**"
    
    return result

def get_user_request_by_class(user_id: int, building: str, class_name: str, date: str = None) -> dict:
    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.execute("""
        SELECT category, quantity
        FROM meals 
        WHERE teacher_id = ? AND date = ? AND building = ? AND class_name = ? AND stage NOT IN ('after_school')
        ORDER BY category
    """, (user_id, date, building, class_name))
    results = cursor.fetchall()
    conn.close()
    
    categories = {}
    for row in results:
        categories[row[0]] = row[1]
    
    return {"categories": categories, "total": sum(categories.values())}

def update_user_request(user_id: int, building: str, class_name: str, new_categories: dict, teacher_name: str) -> bool:
    today = datetime.now().strftime("%Y-%m-%d")
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.execute("""
        SELECT stage, grade, litera FROM meals 
        WHERE teacher_id = ? AND date = ? AND building = ? AND class_name = ?
        LIMIT 1
    """, (user_id, today, building, class_name))
    old_data = cursor.fetchone()
    
    if not old_data:
        conn.close()
        return False
    
    stage, grade, litera = old_data
    
    conn.execute("""
        DELETE FROM meals 
        WHERE teacher_id = ? AND date = ? AND building = ? AND class_name = ?
    """, (user_id, today, building, class_name))
    
    for category, quantity in new_categories.items():
        if quantity > 0:
            conn.execute("""
                INSERT INTO meals 
                (date, building, stage, grade, litera, class_name, category, quantity, teacher_name, teacher_id, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (today, building, stage, grade, litera, class_name, category, quantity, teacher_name, user_id, datetime.now().isoformat()))
    
    conn.commit()
    conn.close()
    return True

def update_user_request_on_date(user_id: int, building: str, class_name: str, new_categories: dict, teacher_name: str, date: str) -> bool:
    """Обновляет заявку на конкретную дату"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.execute("""
        SELECT stage, grade, litera FROM meals 
        WHERE teacher_id = ? AND date = ? AND building = ? AND class_name = ?
        LIMIT 1
    """, (user_id, date, building, class_name))
    old_data = cursor.fetchone()
    
    if not old_data:
        conn.close()
        return False
    
    stage, grade, litera = old_data
    
    # Удаляем старую заявку
    conn.execute("""
        DELETE FROM meals 
        WHERE teacher_id = ? AND date = ? AND building = ? AND class_name = ?
    """, (user_id, date, building, class_name))
    
    # Добавляем новую
    for category, quantity in new_categories.items():
        if quantity > 0:
            conn.execute("""
                INSERT INTO meals 
                (date, building, stage, grade, litera, class_name, category, quantity, teacher_name, teacher_id, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (date, building, stage, grade, litera, class_name, category, quantity, teacher_name, user_id, datetime.now().isoformat()))
    
    conn.commit()
    conn.close()
    return True

def delete_user_request(user_id: int, building: str, class_name: str, date: str = None) -> bool:
    """Удаляет заявку (если дата не указана - сегодняшнюю)"""
    if date is None:
        date = datetime.now().strftime("%Y-%m-%d")
    
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        DELETE FROM meals 
        WHERE teacher_id = ? AND date = ? AND building = ? AND class_name = ?
    """, (user_id, date, building, class_name))
    conn.commit()
    conn.close()
    return True

def delete_user_request_by_date(user_id: int, building: str, class_name: str, date: str) -> bool:
    """Удаляет заявку на конкретную дату"""
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        DELETE FROM meals 
        WHERE teacher_id = ? AND date = ? AND building = ? AND class_name = ?
    """, (user_id, date, building, class_name))
    conn.commit()
    conn.close()
    return True

def get_requests_for_last_days(building: str, days: int = 3) -> list:
    """Получает заявки за последние N дней для администратора"""
    date_to = datetime.now().strftime("%Y-%m-%d")
    date_from = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.execute("""
        SELECT DISTINCT date, class_name, teacher_name, teacher_id, building
        FROM meals 
        WHERE building = ? AND date >= ? AND date <= ?
        ORDER BY date DESC, class_name
    """, (building, date_from, date_to))
    results = cursor.fetchall()
    conn.close()
    
    requests = []
    for row in results:
        requests.append({
            "date": row[0],
            "class_name": row[1],
            "teacher_name": row[2],
            "teacher_id": row[3],
            "building": row[4]
        })
    return requests

def get_request_by_date(building: str, class_name: str, date: str) -> dict:
    """Получает заявку на конкретную дату"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.execute("""
        SELECT category, quantity
        FROM meals 
        WHERE building = ? AND class_name = ? AND date = ?
        ORDER BY category
    """, (building, class_name, date))
    results = cursor.fetchall()
    conn.close()
    
    categories = {}
    for row in results:
        categories[row[0]] = row[1]
    
    return {"categories": categories, "total": sum(categories.values())}

def can_edit_request(date_str: str) -> bool:
    """Проверяет, можно ли редактировать заявку (последние 3 дня)"""
    request_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    today = datetime.now().date()
    delta = (today - request_date).days
    return 0 <= delta <= 3

def get_all_classes_with_requests(building: str, month: int = None, year: int = None) -> list:
    if month is None:
        month = datetime.now().month
    if year is None:
        year = datetime.now().year
    
    first_day = datetime(year, month, 1).strftime("%Y-%m-%d")
    if month == 12:
        last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1) - timedelta(days=1)
    last_day = last_day.strftime("%Y-%m-%d")
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.execute("""
        SELECT DISTINCT class_name
        FROM meals 
        WHERE building = ? AND date >= ? AND date <= ? AND stage NOT IN ('after_school', 'home')
        ORDER BY class_name
    """, (building, first_day, last_day))
    results = cursor.fetchall()
    conn.close()
    
    return [row[0] for row in results]

def get_all_requests_by_class(building: str, class_name: str, month: int = None, year: int = None) -> dict:
    if month is None:
        month = datetime.now().month
    if year is None:
        year = datetime.now().year
    
    first_day = datetime(year, month, 1).strftime("%Y-%m-%d")
    if month == 12:
        last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1) - timedelta(days=1)
    last_day = last_day.strftime("%Y-%m-%d")
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.execute("""
        SELECT date, teacher_name, category, quantity
        FROM meals 
        WHERE building = ? AND class_name = ? AND date >= ? AND date <= ? AND stage NOT IN ('after_school', 'home')
        ORDER BY date, category
    """, (building, class_name, first_day, last_day))
    results = cursor.fetchall()
    conn.close()
    
    by_date = {}
    for row in results:
        date = row[0]
        teacher = row[1]
        category = row[2]
        quantity = row[3]
        
        if date not in by_date:
            by_date[date] = {"teacher": teacher, "categories": {}}
        if category not in by_date[date]["categories"]:
            by_date[date]["categories"][category] = 0
        by_date[date]["categories"][category] += quantity
    
    return by_date

def format_all_requests_by_class(building: str, class_name: str, requests_data: dict, month: int, year: int) -> str:
    if not requests_data:
        return f"📭 Нет заявок для класса {class_name} за {month:02d}.{year}"
    
    month_name = datetime(year, month, 1).strftime("%B %Y")
    result = f"📋 **Все заявки по классу {class_name}**\n"
    result += f"🏫 **Здание:** {building}\n"
    result += f"📅 **Период:** {month_name}\n\n"
    
    total_all = 0
    for date in sorted(requests_data.keys()):
        date_display = datetime.strptime(date, "%Y-%m-%d").strftime("%d.%m.%Y")
        data = requests_data[date]
        teacher = data["teacher"]
        categories = data["categories"]
        
        result += f"**📅 {date_display}**\n"
        result += f"👤 {teacher}\n"
        
        day_total = 0
        for category, quantity in categories.items():
            result += f"   • {category}: {quantity} чел.\n"
            day_total += quantity
        
        result += f"   *Итого за день: {day_total} чел.*\n\n"
        total_all += day_total
    
    result += f"---\n🍽️ **ВСЕГО за месяц: {total_all} чел.**"
    
    return result

def get_user_requests_by_month(user_id: int, month: int, year: int) -> dict:
    first_day = datetime(year, month, 1).strftime("%Y-%m-%d")
    if month == 12:
        last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1) - timedelta(days=1)
    last_day = last_day.strftime("%Y-%m-%d")
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.execute("""
        SELECT date, building, class_name, category, quantity, stage
        FROM meals 
        WHERE teacher_id = ? AND date >= ? AND date <= ?
        ORDER BY date, class_name, category
    """, (user_id, first_day, last_day))
    results = cursor.fetchall()
    conn.close()
    
    by_date = {}
    for row in results:
        date = row[0]
        building = row[1]
        class_name = row[2]
        category = row[3]
        quantity = row[4]
        stage = row[5]
        
        if date not in by_date:
            by_date[date] = {"building": building, "class_name": class_name, "stage": stage, "categories": {}}
        if category not in by_date[date]["categories"]:
            by_date[date]["categories"][category] = 0
        by_date[date]["categories"][category] += quantity
    
    return by_date

def format_user_requests_by_month(requests_data: dict, month: int, year: int, user_name: str) -> str:
    if not requests_data:
        return f"📭 У вас нет заявок за {datetime(year, month, 1).strftime('%B %Y')}"
    
    month_name = datetime(year, month, 1).strftime("%B %Y")
    result = f"📋 **Ваши заявки за {month_name}**\n👤 {user_name}\n\n"
    
    total_all = 0
    for date in sorted(requests_data.keys()):
        date_display = datetime.strptime(date, "%Y-%m-%d").strftime("%d.%m.%Y")
        data = requests_data[date]
        
        stage_display = ""
        if data["stage"] == "home":
            stage_display = " 🏠 (Надомное)"
        elif data["stage"] == "after_school":
            stage_display = " ⏰ (Продленка)"
        
        result += f"**📅 {date_display}**\n"
        result += f"🏫 {data['building']} | 📖 {data['class_name']}{stage_display}\n"
        
        day_total = 0
        for category, quantity in data["categories"].items():
            result += f"   • {category}: {quantity} чел.\n"
            day_total += quantity
        
        result += f"   *Итого за день: {day_total} чел.*\n\n"
        total_all += day_total
    
    result += f"---\n🍽️ **ВСЕГО за месяц: {total_all} чел.**"
    
    return result

def create_excel_report_for_building(building: str, date_from: str, date_to: str, period_type: str) -> str:
    df = get_report_by_building_and_date_range(building, date_from, date_to)
    
    building_folder = os.path.join(REPORTS_DIR, building)
    os.makedirs(building_folder, exist_ok=True)
    
    now = datetime.now()
    date_from_display = datetime.strptime(date_from, "%Y-%m-%d").strftime("%d.%m.%Y")
    date_to_display = datetime.strptime(date_to, "%Y-%m-%d").strftime("%d.%m.%Y")
    
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    
    if period_type == "daily":
        filename = os.path.join(building_folder, f"report_{building}_daily_{date_from}_{timestamp}.xlsx")
        sheet_title = f"отчёт за {date_from_display}"
    elif period_type == "weekly":
        filename = os.path.join(building_folder, f"report_{building}_weekly_{date_from}_to_{date_to}_{timestamp}.xlsx")
        sheet_title = f"отчёт за период {date_from_display} - {date_to_display}"
    else:
        month_name = now.strftime("%B %Y")
        filename = os.path.join(building_folder, f"report_{building}_monthly_{now.strftime('%Y_%m')}_{timestamp}.xlsx")
        sheet_title = f"отчёт за {month_name}"
    
    categories_1_4 = ["1-4 класс", "ОВЗ и инвалиды 1-4 класс"]
    categories_5_11 = [
        "Без субсидии", "Малообеспеченные", "Многодетные",
        "Участники боевых действий", "Семьи в ТЖС", "С нарушениями здоровья",
        "Семьи военнослужащих", "ОВЗ и инвалиды 5-11", "Кадетские классы"
    ]
    
    classes_by_stage = {
        "1": [f"{g}.{l}" for g in ["1", "2", "3", "4"] for l in range(1, 10)],
        "2": [f"{g}.{l}" for g in ["5", "6", "7", "8", "9"] for l in range(1, 10)],
        "3": ["10.1", "10.2", "11.1", "11.2"]
    }
    
    stage_info = {
        "1": {"sheet_name": "1-4 классы", "categories": categories_1_4, "all_classes": classes_by_stage["1"]},
        "2": {"sheet_name": "5-9 классы" if building == "Танкистов" else "5-11 классы", "categories": categories_5_11, "all_classes": classes_by_stage["2"]},
        "3": {"sheet_name": "10-11 классы", "categories": categories_5_11, "all_classes": classes_by_stage["3"]}
    }
    
    # Для Танкистов убираем 10-11 классы
    if building == "Танкистов":
        del stage_info["3"]
        stage_info["2"]["sheet_name"] = "5-9 классы"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Основные листы по ступеням
        for stage_code, info in stage_info.items():
            sheet_name = info["sheet_name"]
            all_categories = info["categories"]
            all_classes = info["all_classes"]
            
            stage_df = df[df['stage'] == stage_code] if not df.empty else pd.DataFrame()
            
            data_dict = {}
            if not stage_df.empty:
                for _, row in stage_df.iterrows():
                    class_name = row['class_name']
                    category = row['category']
                    quantity = row['quantity']
                    if class_name not in data_dict:
                        data_dict[class_name] = {}
                    data_dict[class_name][category] = quantity
            
            data = []
            for category in all_categories:
                row = {"Категория": category}
                for class_name in all_classes:
                    if class_name in data_dict and category in data_dict[class_name]:
                        row[class_name] = data_dict[class_name][category]
                    else:
                        row[class_name] = 0
                data.append(row)
            
            existing_classes = [c for c in all_classes if any(row[c] > 0 for row in data)]
            if not existing_classes and not stage_df.empty:
                existing_classes = all_classes[:5]
            if not existing_classes:
                existing_classes = all_classes[:5]
            
            filtered_data = []
            for row in data:
                filtered_row = {"Категория": row["Категория"]}
                for class_name in existing_classes:
                    filtered_row[class_name] = row[class_name]
                filtered_row["ИТОГО по категории"] = sum(row[class_name] for class_name in existing_classes)
                filtered_data.append(filtered_row)
            
            result_df = pd.DataFrame(filtered_data)
            
            totals_by_class = {"Категория": "ВСЕГО по классу"}
            grand_total = 0
            for class_name in existing_classes:
                class_total = result_df[class_name].sum()
                totals_by_class[class_name] = class_total
                grand_total += class_total
            totals_by_class["ИТОГО по категории"] = grand_total
            result_df = pd.concat([result_df, pd.DataFrame([totals_by_class])], ignore_index=True)
            
            result_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)
            
            worksheet = writer.sheets[sheet_name]
            
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(existing_classes)+2)
            worksheet.cell(row=1, column=1, value=sheet_title)
            worksheet.cell(row=1, column=1).font = Font(size=14, bold=True)
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
            
            for col, class_name in enumerate(existing_classes, 2):
                cell = worksheet.cell(row=3, column=col, value=class_name)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            cell = worksheet.cell(row=3, column=len(existing_classes)+2, value="ИТОГО по категории")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            worksheet.cell(row=3, column=1, value="Категория")
            worksheet.cell(row=3, column=1).font = Font(bold=True)
            worksheet.cell(row=3, column=1).alignment = Alignment(horizontal='center')
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            max_row = worksheet.max_row
            max_col = len(existing_classes) + 2
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    if row >= 3:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                if row >= 4:
                    worksheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            
            for col in range(1, max_col + 1):
                max_length = 0
                for row in range(1, max_row + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max_length + 2, 35)
                worksheet.column_dimensions[get_column_letter(col)].width = adjusted_width
            
            worksheet.column_dimensions['A'].width = 30
        
        # ========== ЛИСТ ДЛЯ НАДОМНОГО ОТДЕЛЕНИЯ (только для Марченко) ==========
        if building == "Марченко":
            print("🔍 Формирую лист Надомное отделение...")
            home_df = get_home_requests(date_from, date_to)
            
            if home_df.empty:
                # Создаём пустую структуру с классами
                classes = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11"]
                categories = categories_5_11
                
                data = []
                for category in categories:
                    row = {"Категория": category}
                    for class_name in classes:
                        row[class_name] = 0
                    data.append(row)
                
                result_df = pd.DataFrame(data)
                
                totals_by_class = {"Категория": "ВСЕГО по классу"}
                for class_name in classes:
                    totals_by_class[class_name] = 0
                totals_by_class["ИТОГО по категории"] = 0
                result_df = pd.concat([result_df, pd.DataFrame([totals_by_class])], ignore_index=True)
            else:
                classes = sorted(home_df["class_name"].unique())
                categories = categories_5_11
                
                data = []
                for category in categories:
                    row = {"Категория": category}
                    for class_name in classes:
                        subset = home_df[(home_df["category"] == category) & (home_df["class_name"] == class_name)]
                        row[class_name] = subset["quantity"].sum() if not subset.empty else 0
                    data.append(row)
                
                result_df = pd.DataFrame(data)
                
                totals_by_class = {"Категория": "ВСЕГО по классу"}
                grand_total = 0
                for class_name in classes:
                    class_total = result_df[class_name].sum()
                    totals_by_class[class_name] = class_total
                    grand_total += class_total
                totals_by_class["ИТОГО по категории"] = grand_total
                result_df = pd.concat([result_df, pd.DataFrame([totals_by_class])], ignore_index=True)
            
            result_df.to_excel(writer, sheet_name="Надомное отделение", index=False, startrow=3)
            
            worksheet = writer.sheets["Надомное отделение"]
            
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(classes)+2)
            worksheet.cell(row=1, column=1, value=f"Надомное отделение - {sheet_title}")
            worksheet.cell(row=1, column=1).font = Font(size=14, bold=True)
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
            
            for col, class_name in enumerate(classes, 2):
                cell = worksheet.cell(row=3, column=col, value=class_name)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            cell = worksheet.cell(row=3, column=len(classes)+2, value="ИТОГО по категории")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            worksheet.cell(row=3, column=1, value="Категория")
            worksheet.cell(row=3, column=1).font = Font(bold=True)
            worksheet.cell(row=3, column=1).alignment = Alignment(horizontal='center')
            
            max_row = worksheet.max_row
            max_col = len(classes) + 2
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    if row >= 3:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                if row >= 4:
                    worksheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            
            for col in range(1, max_col + 1):
                max_length = 0
                for row in range(1, max_row + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max_length + 2, 35)
                worksheet.column_dimensions[get_column_letter(col)].width = adjusted_width
            worksheet.column_dimensions['A'].width = 30
            print("✅ Лист Надомное отделение добавлен")
        
        # ========== ЛИСТ ДЛЯ ПРОДЛЕНКИ ==========
        print(f"🔍 Формирую лист Продленка для {building}...")
        after_school_df = get_after_school_requests(building, date_from, date_to)
        
        if not after_school_df.empty:
            # Группируем по датам
            by_date = {}
            for _, row in after_school_df.iterrows():
                date = row['date']
                class_name = row['class_name']
                quantity = row['quantity']
                if date not in by_date:
                    by_date[date] = {}
                by_date[date][class_name] = quantity
            
            data = []
            all_classes = set()
            for date, classes in by_date.items():
                all_classes.update(classes.keys())
            all_classes = sorted(all_classes)
            
            for date, classes in sorted(by_date.items()):
                date_display = datetime.strptime(date, "%Y-%m-%d").strftime("%d.%m.%Y")
                row = {"Дата": date_display}
                for class_name in all_classes:
                    row[class_name] = classes.get(class_name, 0)
                row["ИТОГО за день"] = sum(classes.values())
                data.append(row)
            
            result_df = pd.DataFrame(data)
            totals_row = {"Дата": "ВСЕГО"}
            for class_name in all_classes:
                totals_row[class_name] = result_df[class_name].sum()
            totals_row["ИТОГО за день"] = result_df["ИТОГО за день"].sum()
            result_df = pd.concat([result_df, pd.DataFrame([totals_row])], ignore_index=True)
            result_df.to_excel(writer, sheet_name="Продленка", index=False, startrow=3)
            
            worksheet = writer.sheets["Продленка"]
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_classes)+2)
            worksheet.cell(row=1, column=1, value=f"Продленка - {sheet_title}")
            worksheet.cell(row=1, column=1).font = Font(size=14, bold=True)
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
            for col, class_name in enumerate(all_classes, 2):
                cell = worksheet.cell(row=3, column=col, value=class_name)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            cell = worksheet.cell(row=3, column=len(all_classes)+2, value="ИТОГО за день")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            worksheet.cell(row=3, column=1, value="Дата")
            worksheet.cell(row=3, column=1).font = Font(bold=True)
            worksheet.cell(row=3, column=1).alignment = Alignment(horizontal='center')
            max_row = worksheet.max_row
            max_col = len(all_classes) + 2
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    if row >= 3:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                if row >= 4:
                    worksheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            for col in range(1, max_col + 1):
                max_length = 0
                for row in range(1, max_row + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max_length + 2, 25)
                worksheet.column_dimensions[get_column_letter(col)].width = adjusted_width
            print("✅ Лист Продленка добавлен")
        else:
            print(f"⚠️ Нет данных для Продленки в здании {building}")
    
    print(f"✅ Отчёт создан: {os.path.abspath(filename)}")
    
    try:
        from email_sender import send_report_via_email
        send_report_via_email(filename, period_type, building)
        print(f"📧 Отчёт отправлен на email")
    except Exception as e:
        print(f"❌ Ошибка отправки email: {e}")
    
    return filename

init_db()
print("✅ storage_food.py загружен")
